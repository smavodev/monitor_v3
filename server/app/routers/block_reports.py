from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import datetime
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.db import get_db
from core.permissions import require_permission
from models.models import Agent, Sede, BlockedSite
from routers.blocked_sites import resolve_domains_for_agent
from routers.block_schedules import resolve_schedule_for_agent, WEEKDAY_KEYS, _local_now

router = APIRouter(prefix="/api/blocked-sites", tags=["blocked-sites-report"])

WEEKDAY_LABELS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

FILL_HEADER  = PatternFill("solid", fgColor="1E2540")
FILL_BLOCKED = PatternFill("solid", fgColor="FDE2E2")
FILL_FREE    = PatternFill("solid", fgColor="E3F5E9")
FONT_HEADER  = Font(bold=True, color="FFFFFF")
WRAP_TOP     = Alignment(wrap_text=True, vertical="top")

def _build_report(db: Session, sede_id: Optional[str], agent_id: Optional[str], only_with_blocks: bool) -> io.BytesIO:
    q = db.query(Agent)
    if sede_id:
        q = q.filter(Agent.sede_id == sede_id)
    if agent_id:
        q = q.filter(Agent.id == agent_id)
    agents = q.order_by(Agent.hostname).all()

    active_sites = db.query(BlockedSite).filter(BlockedSite.active == True).all()
    sedes_by_id = {s.id: s for s in db.query(Sede).all()}
    now = _local_now(None)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de bloqueos"

    headers = ["Equipo", "Área", "Sitios bloqueados", "Excepciones (motivo / expira)",
               "Horario aplicado", "Horario expira"] + WEEKDAY_LABELS
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_idx = 2
    for agent in agents:
        domains = sorted(resolve_domains_for_agent(agent, active_sites))
        if only_with_blocks and not domains:
            continue

        exceptions = [
            s for s in active_sites
            if s.agent_id == agent.id and s.is_exception and (s.expires_at is None or s.expires_at > now)
        ]
        exc_text = "\n".join(
            s.domain + (f" — {s.reason}" if s.reason else "") +
            (f" (expira {s.expires_at.strftime('%Y-%m-%d')})" if s.expires_at else "")
            for s in exceptions
        ) or "—"

        schedule = resolve_schedule_for_agent(agent, db)
        area_name = sedes_by_id[agent.sede_id].name if agent.sede_id in sedes_by_id else "Sin área"

        if schedule is None:
            origen, expira_txt, days = "Sin horario configurado", "—", {}
        elif schedule.agent_id == agent.id:
            origen = "Propio (excepción de equipo)"
            expira_txt = schedule.expires_at.strftime("%Y-%m-%d") if schedule.expires_at else "—"
            days = schedule.days or {}
        elif schedule.sede_id:
            origen, expira_txt, days = f"Heredado del área: {area_name}", "—", (schedule.days or {})
        else:
            origen, expira_txt, days = "Horario global", "—", (schedule.days or {})

        row = [
            agent.display_name or agent.hostname, area_name,
            "\n".join(domains) if domains else "—", exc_text, origen, expira_txt,
        ]
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = WRAP_TOP

        for i, key in enumerate(WEEKDAY_KEYS):
            col = 7 + i
            rng = days.get(key)
            cell = ws.cell(row=row_idx, column=col)
            if rng:
                cell.value = f"{rng[0]}-{rng[1]}"
                cell.fill = FILL_BLOCKED
            else:
                cell.value = "Libre"
                cell.fill = FILL_FREE
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row_idx += 1

    widths = [22, 16, 34, 34, 26, 14] + [11] * 7
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

@router.get("/report.xlsx")
def download_report(
    sede_id: Optional[str] = Query(None),
    agent_id: Optional[str] = Query(None),
    only_with_blocks: bool = Query(False),
    user=Depends(require_permission("parental_blocked", "view")), db: Session = Depends(get_db),
):
    buf = _build_report(db, sede_id or None, agent_id or None, only_with_blocks)
    filename = f"smartmonitor_bloqueos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
